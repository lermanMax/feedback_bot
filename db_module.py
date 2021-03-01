import psycopg2
from psycopg2.extras import DictCursor
from psycopg2.sql import SQL, Identifier

import datetime 

import openpyxl
from openpyxl.styles import Font

class DB_module:
    
    def __init__(self, DB_HOST, DB_NAME, DB_USER, DB_PASS, DB_PORT):
        self.DB_dict = {'host': DB_HOST,
                        'database': DB_NAME,
                        'user': DB_USER,
                        'password': DB_PASS,
                        'port': DB_PORT}

    """users"""    
    def add_user(self,user_id,first_name, last_name, username, language_code):
        if self.get_user(user_id): return 'user exist'
        
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
    
        insert_into_users="""insert into users
                          (id, first_name, last_name, username, language_code) 
                          values (%s, %s, %s, %s, %s);"""
        values = (user_id, first_name, last_name, username, language_code)
                
        cur.execute(insert_into_users, values)
        conn.commit()
        cur.close()
        conn.close()
        return 'user added'
            
    def add_phone(self, user_id = None, phone = None):
        if not self.get_user(user_id): return 'user not exist'
        
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
                
        update_users = """UPDATE users SET phone = %s
                        WHERE id = %s ;""" 
        data = (phone, user_id)
        
        cur.execute(update_users, data)
        conn.commit()
        cur.close()
        conn.close()
        return 'phone added'
                    
    def get_user(self, user_id):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
                
        select_users = "select * from users where id = %s;"
        data = (user_id,)
        
        cur.execute(select_users, data)
        result = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    """menu"""
    def add_menu(self, date=None, str_date=None): #str_date = '2021.01.01'
        menu_id = 0
            
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
        
        if str_date:
            date_list = str_date.split('.')
            date_list = [int(i) for i in date_list]
            date = datetime.date(date_list[0], date_list[1], date_list[2])
                
        insert_into_menu="""insert into menu (date) values (%s);"""
        values = (date,)
        
        if date <= datetime.date.today(): 
            # 'Ошибка. Можно добавлять только меню будущих дней'
            menu_id = None
        else:
            menu = self.get_menu(date=date) 
            if not menu:
                cur.execute(insert_into_menu, values)
                conn.commit()
                menu_id = self.get_menu(date=date)['id']
            else:
                menu_id = menu['id']
                delete_from_dish = 'delete from dish where menu_id = (%s);'
                cur.execute(delete_from_dish, (menu_id,))
                conn.commit()
                
        conn.commit()
        cur.close()
        conn.close()
        return menu_id
    
    def get_menu(self, menu_id=None, date=None, keyword=None):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
        
        if date:        
            select_menu = "select * from menu where date = %s;"
            data = (date,)
        elif keyword:
            today = datetime.date.today()
            one_day_ago = today - datetime.timedelta(days=1)
            two_days_ago = today - datetime.timedelta(days=2)
    
            key_date = {'Меню позавчера': two_days_ago,
                        'Меню вчера': one_day_ago,
                        'Меню сегодня': datetime.date.today()}
            
            select_menu = "select * from menu where date = %s;"
            data = (key_date[keyword],)
        elif menu_id:
            select_menu = "select * from menu where id = %s;"
            data = (menu_id,)
        
        cur.execute(select_menu, data)
        result = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    """dish"""
    def add_dish(self, menu_id, name):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
                
        insert_into_dish="""insert into dish (menu_id, name) values (%s, %s);"""
        values = (menu_id, name)
        
        cur.execute(insert_into_dish, values)
        
        conn.commit()
        cur.close()
        conn.close()
        return 'dish added'
    
    def get_dish(self, dish_id=None, menu_id=None):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
        
        if dish_id:        
            select_dish = "select * from dish where id = %s;"
            data = (dish_id,)
            cur.execute(select_dish, data)
            result = cur.fetchone()
        elif menu_id:
            select_dish = "select * from dish where menu_id = %s;"
            data = (menu_id,)
            cur.execute(select_dish, data)
            result = cur.fetchall()
        
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    """feedback"""
    def add_feedback(self, user_id, dish_id, mark=0, review=" "):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
        
        if review != " ": review += '; '
        
        date = datetime.date.today()
        insert_into_dish="""insert into feedback 
                            (user_id, dish_id, date, mark, review) 
                            values (%s, %s, %s, %s, %s);"""
        values = (user_id, dish_id, date, mark, review)
        
        cur.execute(insert_into_dish, values)
        conn.commit()
        cur.close()
        conn.close()
        return 'feedback added'
    
    def update_feedback(self, user_id, dish_id, mark=None, review=" "):
        if self.get_feedback(user_id, dish_id):    
            conn = psycopg2.connect(**self.DB_dict)
            cur = conn.cursor(cursor_factory = DictCursor)
            
            if mark:
                update_feedback_= """UPDATE feedback SET mark = %s
                                    WHERE user_id = %s and dish_id = %s;"""
                values = (mark, user_id, dish_id)
            elif review:
                update_feedback_= """UPDATE feedback 
                                    SET review = review || %s
                                    WHERE user_id = %s and dish_id = %s;"""
                values = (review+'; ', user_id, dish_id)
            else:
                print('no mark, no review')
                update_feedback_= None
                values = None
                
            cur.execute(update_feedback_, values)
            conn.commit()
            cur.close()
            conn.close()
        else:
            self.add_feedback(user_id, dish_id, mark, review)
        return 'feedback updated'
    
    def get_feedback(self, user_id, dish_id):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
               
        select_dish = "select * from feedback where user_id = %s and dish_id = %s ;"
        data = (user_id, dish_id)
        cur.execute(select_dish, data)
        result = cur.fetchone()
        
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    """answer_loyalty"""
    def get_answer(self, answer_id):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
        
        select_answer_loyalty = "select * from answer_loyalty where id = %s ;"
        data = (answer_id,)
           
        cur.execute(select_answer_loyalty, data)
        result = cur.fetchone()
        
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    def get_answers(self, user_id = None, answering_date = None):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
                    
        if (user_id != None):
            select_answer_loyalty = "select * from answer_loyalty where user_id = %s ;"
            data = (user_id,)
        elif (answering_date != None):
            select_answer_loyalty = "select * from answer_loyalty where answering_date = %s ;"
            data = (answering_date,)
        else:
            select_answer_loyalty = "select * from answer_loyalty ;"
            data = ()
            
        cur.execute(select_answer_loyalty, data)
        result = cur.fetchall()
        
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    def get_last_answer(self, user_id):
        conn = psycopg2.connect(**self.DB_dict)
        cur = conn.cursor(cursor_factory = DictCursor)
        
        lines = self.get_answers(user_id = user_id)
        max_id = 0
        for ans in lines: max_id = max(int(ans['id']), max_id)
        
        result = self.get_answer(answer_id = max_id)
        
        conn.commit()
        cur.close()
        conn.close()
        return result
    
    def add_answer(self, user_id = None, question_name = None, answer = None):
        
        first_question_name = 'loyalty'
    
        with psycopg2.connect(**self.DB_dict) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                if question_name != first_question_name: 
                    cur.execute(f"select * from answer_loyalty where user_id = {user_id};")
                    lines = cur.fetchall()
                    max_id = 0
                    for ans in lines: max_id = max(int(ans['id']), max_id)
                    
                    cur.execute(SQL("""
                                UPDATE answer_loyalty SET 
                                {} = %(answer)s
                                WHERE
                                id = %(id)s
                                ;
                                """).format(Identifier(question_name)), 
                                {
                                'id': max_id,
                                'answer': answer
                                })
                    
                
                else: # 
                    answering_date = datetime.date.today()
                    cur.execute("""
                                insert into answer_loyalty
                                (user_id, answering_date, loyalty, review) 
                                values
                                (%(user_id)s, %(answering_date)s, %(loyalty)s, '-');
                                """, 
                                {
                                 'user_id': user_id, 
                                 'answering_date': answering_date, 
                                 'loyalty': answer,
                                 })
                return 
            
    def add_review(self, user_id = None, text = None):
    
        with psycopg2.connect(**self.DB_dict) as conn:            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                cur.execute(f"select * from answer_loyalty where user_id = {user_id};")
                lines = cur.fetchall()
                max_id = 0
                for ans in lines: max_id = max(int(ans['id']), max_id)
                
                cur.execute("""
                                UPDATE answer_loyalty SET 
                                review = review || %(text)s
                                WHERE
                                id = %(id)s
                                ;
                                """, 
                                {
                                'id': max_id,
                                'text': text+'\n'
                                })
    
    """anther"""
    def export_answer_loyalty_to_excel(self, headings, filepath):

        """
        Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.
    
        Arguments:
        connection - an open psycopg2 (this function does not close the connection)
        query_string - SQL to get data
        headings - list of strings to use as column headings
        filepath - path and filename of the Excel file
    
        psycopg2 and file handling errors bubble up to calling code.
        """
        
        with psycopg2.connect(**self.DB_dict) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute("""
                            select u.*, al.* from answer_loyalty al 
                            left join users u on u.id = al.user_id;
                            """)
                data = cur.fetchall()
    
    
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title='Все ответы', index=0)
    
        sheet.row_dimensions[1].font = Font(bold = True)
    
        # Spreadsheet row and column indexes start at 1
        # so we use "start = 1" in enumerate so
        # we don't need to add 1 to the indexes.
        for colno, heading in enumerate(headings, start = 1):
            sheet.cell(row = 1, column = colno).value = heading
    
        # This time we use "start = 2" to skip the heading row.
        for rowno, row in enumerate(data, start = 2):
            for colno, cell_value in enumerate(row, start = 1):
                sheet.cell(row = rowno, column = colno).value = cell_value
    
        wb.save(filepath) 
        
        
    def export_menu_feedback_to_excel(self, headings, filepath):

        """
        Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.
    
        Arguments:
        connection - an open psycopg2 (this function does not close the connection)
        query_string - SQL to get data
        headings - list of strings to use as column headings
        filepath - path and filename of the Excel file
    
        psycopg2 and file handling errors bubble up to calling code.
        """
        
        with psycopg2.connect(**self.DB_dict) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute("""
                            select 
                            first_name, last_name, username,
                            phone, name, m.date, fb.date, mark, review
                            from feedback fb 
                            JOIN users u on u.id = fb.user_id 
                            JOIN dish d  on d.id = fb.dish_id
                            JOIN menu m on d.menu_id = m.id;
                            """)
                data = cur.fetchall()
    
    
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title='Оценка меню', index=0)
    
        sheet.row_dimensions[1].font = Font(bold = True)
    
        # Spreadsheet row and column indexes start at 1
        # so we use "start = 1" in enumerate so
        # we don't need to add 1 to the indexes.
        for colno, heading in enumerate(headings, start = 1):
            sheet.cell(row = 1, column = colno).value = heading
    
        # This time we use "start = 2" to skip the heading row.
        for rowno, row in enumerate(data, start = 2):
            for colno, cell_value in enumerate(row, start = 1):
                sheet.cell(row = rowno, column = colno).value = cell_value
    
        wb.save(filepath)        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        